import openpyxl

#def func(x,y):
    
#    wb = openpyxl.load_workbook('Table.xlsx')
    
#    sheetNames = wb.get_sheet_names()
    
#func(sys.argv[1], sys.argv[2])

wb = openpyxl.load_workbook('Table.xlsx', data_only=True)
sheetNames = wb.get_sheet_names()
parsingError = 0
retLine = ""

for i in range(0,len(sheetNames)):
    rezultLine = "|"                             #буфферная переменная для выведения результата
    sheet = wb.get_sheet_by_name(sheetNames[i])
    if (sheet['A1'].value == "Yes"):            #проверили, что страницу надо анализировать
        rezultLine += sheetNames[i]
        if type(sheet['F3'].value) is int and type(sheet['F4'].value) is int:   #проверяем на значение int
            rezultLine += ":" + str(sheet['F3'].value) + ":" + str(sheet['F4'].value) + ":"
            for k in range(2,sheet.max_row):                 #начинаем со второй строки ибо первая это Name
                if str(sheet.cell(row = k,column = 20).value) != "None" and str(sheet.cell(row = k,column = 20).value) != " ":      #если ячейка элемента пуста, то элемента нет
                    rezultLine += "/" + str(sheet.cell(row = k,column = 19).value) + "~" + str(sheet.cell(row = k,column = 20).value) #19 - это порядковый номер буквы S по горизонтали (номера элементов) 20 - это порядковый номер буквы Т по горизонтали (имена элементов)
        else:
            rezultLine += ":parsingError (value not int type)"
            parsingError += 1
        retLine += rezultLine



retLine += "|parsingErrorCounter=" + str(parsingError)
print(retLine)




#print(sheetNames)
#sheet = wb.get_sheet_by_name('DevCtrl')
#print(sheet['T2'].value)
